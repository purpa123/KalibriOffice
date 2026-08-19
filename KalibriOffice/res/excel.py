# C:\Users\georg\Projects\CalibriOffice\res\excel.py

import customtkinter as ctk
from tkinter import filedialog, messagebox
import os
import sys

try:
    from tksheet import Sheet
    TKSHEET_AVAILABLE = True
except ImportError:
    TKSHEET_AVAILABLE = False
    print("⚠️ For tables install: pip install tksheet")

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ For Excel install: pip install openpyxl")

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class KalibriTable(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("KalibriOffice — KalibriTable")
        self.geometry("1100x750")
        self.minsize(800, 500)

        self.current_file = None
        self.data = []

        self.create_menu()
        self.create_table()

        print("📊 KalibriTable started (English UI)")

    def create_menu(self):
        menu_frame = ctk.CTkFrame(self, height=50, fg_color="transparent")
        menu_frame.pack(fill="x", padx=10, pady=(10, 0))

        ctk.CTkButton(menu_frame, text="📂 Open", command=self.open_file, width=80).pack(side="left", padx=2)
        ctk.CTkButton(menu_frame, text="💾 Save", command=self.save_file, width=80).pack(side="left", padx=2)
        ctk.CTkButton(menu_frame, text="📁 Save As", command=self.save_as_file, width=110).pack(side="left", padx=2)
        ctk.CTkButton(menu_frame, text="➕ Add Row", command=self.add_row, width=130).pack(side="left", padx=2)
        ctk.CTkButton(menu_frame, text="➖ Delete Row", command=self.delete_row, width=130).pack(side="left", padx=2)

        self.label_info = ctk.CTkLabel(menu_frame, text="Untitled")
        self.label_info.pack(side="right", padx=10)

    def create_table(self):
        if not TKSHEET_AVAILABLE:
            messagebox.showerror("Error", "tksheet not installed.\nInstall: pip install tksheet")
            return

        self.table_frame = ctk.CTkFrame(self)
        self.table_frame.pack(padx=10, pady=(10, 10), fill="both", expand=True)

        self.sheet = Sheet(
            self.table_frame,
            data=[[]],
            headers=["A", "B", "C", "D", "E"],
            width=900,
            height=500
        )
        self.sheet.pack(fill="both", expand=True)
        self.sheet.bind("<ButtonRelease-1>", self.on_cell_change)

        self.data = [["", "", "", "", ""] for _ in range(10)]
        self.sheet.set_sheet_data(self.data)
        self.sheet.headers(["A", "B", "C", "D", "E"])

    def on_cell_change(self, event):
        try:
            self.data = self.sheet.get_sheet_data()
        except:
            pass

    def get_data(self):
        try:
            return self.sheet.get_sheet_data()
        except:
            return self.data

    def add_row(self):
        data = self.get_data()
        if data:
            num_cols = len(data[0]) if data else 5
            data.append(["" for _ in range(num_cols)])
            self.sheet.set_sheet_data(data)
            self.data = data

    def delete_row(self):
        data = self.get_data()
        if data and len(data) > 1:
            data.pop()
            self.sheet.set_sheet_data(data)
            self.data = data
        else:
            messagebox.showinfo("KalibriTable", "Cannot delete the last row")

    def open_file(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl not installed.\nInstall: pip install openpyxl")
            return

        file_path = filedialog.askopenfilename(
            title="Open file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append([cell if cell is not None else "" for cell in row])
            if not data:
                data = [[""] * 5]
            self.data = data
            self.sheet.set_sheet_data(data)
            if data and len(data[0]) > 0:
                self.sheet.headers([chr(65 + i) for i in range(len(data[0]))])
            self.current_file = file_path
            self.update_title()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file:\n{str(e)}")

    def save_file(self):
        if self.current_file:
            self._save_to_file(self.current_file)
        else:
            self.save_as_file()

    def save_as_file(self):
        file_path = filedialog.asksaveasfilename(
            title="Save file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return
        self._save_to_file(file_path)

    def _save_to_file(self, file_path):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl not installed.")
            return
        try:
            data = self.get_data()
            wb = Workbook()
            ws = wb.active
            for row_idx, row in enumerate(data, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            wb.save(file_path)
            self.current_file = file_path
            self.update_title()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save:\n{str(e)}")

    def update_title(self):
        if self.current_file:
            name = os.path.basename(self.current_file)
            self.title(f"KalibriTable — {name}")
            self.label_info.configure(text=name)
        else:
            self.title("KalibriTable — Untitled")
            self.label_info.configure(text="Untitled")

if __name__ == "__main__":
    app = KalibriTable()
    app.mainloop()